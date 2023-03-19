# Maxy
import datetime
import pyttsx3
import smtplib  ## for email send
import speech_recognition as sr
import webbrowser
import os
import subprocess  ## for Run any application
import random
import math
import wikipedia
import pywhatkit as wtp  ## for youtube video play
import requests  ## for temperature get 
from googlesearch import search
import openpyxl  ## for exel file open and read

# create date today
date = datetime.datetime.now()
date = str(date)
date = date.split()
date = date[0]

# create random number
r = random.randint(0, 7)


# Create a engine
engine = pyttsx3.init()

# set voice of male or female
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id)

# set voice rate
rate = engine.getProperty('rate')
engine.setProperty('rate', 178)

# create current datetime
cur_time = datetime.datetime.now().strftime('%H:%M:%S')

# take only hour
hour = int(datetime.datetime.now().hour)
minut = int(datetime.datetime.now().minute)

# whether


def speak(text):
    engine.say(text)
    engine.runAndWait()


def wishme(text):
    print(text)
    print()
    if hour >= 0 and hour <= 12:
        print('good morning sir,\n I am maxi, \n  how may i help you')
        speak('good morning sir, I am maxi, \n  how may i help you')

    elif hour >= 12 and hour <= 18:
        print('good afternoon sir, \nI am maxi,  how may i help you')
        speak('good afternoon sir, I am maxi,  how may i help you')

    else:
        print('good evening sir,\n I am maxi,  how may i help you')
        speak('good evening sir, I am maxi,  how may i help you')


def sendmail(email,subject,contant,name):
    from email.message import EmailMessage
    email_id = "lalitmaxbusiness@gmail.com"
    email_pas = "iwitlggtpqxppmup"

    msg = EmailMessage()
    msg["subject"] = subject
    msg["from"] = "Lalit kumar yadav"
    msg["to"] = email
    msg.set_content(contant)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(email_id, email_pas)
        smtp.send_message(msg)
        print("sent mail succesfully...")  
        speak("sent mail succesfully")  

def lower(t):
    return t.lower()


#  take command for do anything
def command():
    r = sr.Recognizer()
    with sr.Microphone() as mic:
        print('listening...')
        r.pause_threshold = 2
        audio = r.listen(mic)

    try:
        print('recognizing...')
        text = r.recognize_google(audio)
        text = text.lower()
    except Exception as e:
        print(e)
        print("sorry, i don't understand")
        speak("sorry, i don't understand")
        return 'none'
    return text


def doooo(text):
    if text in ['what is the time', "what's the time", 'is time', 'time now', ]:
        print(text)
        print()
        if hour >= 12:
            print(hour, ":", minut, 'PM')
            speak(str(hour)+":"+str(minut)+'PM')
        else:
            print(hour, ":", minut, 'AM')
            speak(str(hour)+":"+str(minut)+'Am')

    elif "open gmail account" in text:
        rm = openpyxl.load_workbook("Chitkara_email.xlsx")

        sh1 = rm["Sheet1"]

        length = sh1.max_row

        # for read
        li_name = []
        for i in range(1, length+1):
             li_name.append(sh1.cell(i, 1).value)


        speak("which person you want to send email")
        name=command()
        print(name)
        speak("tell me subject")
        subject=command()
        print(subject)
        speak("what do you want to send")
        contant=command()
        print(contant)

        for i in range(1, length):
            ans=lower(li_name[i])
            if name==ans:
                 email=sh1.cell(i+1, 2).value

        sendmail(email,subject,contant,name)


    elif ("weather of") in text or "weather in" in text:
        print(text)
        print()

        t = text.split()

        city_name = t[-1]
        data = requests.get("https://api.openweathermap.org/data/2.5/weather?q=" +
                            city_name+"&appid=882bef32516f8990d82b762cae54604c").json()

        weather = "weather is, "+data["weather"][0]["main"]
        Temp = "temperature is, " +str(int(data["main"]["temp"]-273.5))+" degree Celsius"
        description = 'description is, '+data["weather"][0]['description']
        name = "Name, "+data["name"]+","
        code = "Code is, "+str(data["cod"])

        print(weather)
        print(Temp)
        print(description)
        print(name)
        print(code)
        speak(weather)
        speak(Temp+description)
        speak(name+code)

    elif "temperature in" in text:
        print(text)
        print()
        t = text.split()

        city_name = t[-1]
        data = requests.get("https://api.openweathermap.org/data/2.5/weather?q=" +
                            city_name+"&appid=882bef32516f8990d82b762cae54604c").json()

        Temp = "The temperature is, " + \
            str(int(data["main"]["temp"]-272))+" degree Celsius,"
        name = "in, "+data["name"]
        print(Temp)
        print(name)
        speak(Temp+name)

    elif text in ['send mail to lalit', 'send email to lalit', 'send mail lalit']:
        print(text)
        print(text)
        print('what do you want to write.. ')
        mail = command()
        sendmail(mail)
    elif text in ["exit", "quit"]:
        exit()
    elif text in ["play video", "play song", "video"]:
        webbrowser.open("www.youtube.com")

    elif text in ['open google', 'google open']:
        webbrowser.open('www.google.com')
        speak('opening google')

    elif text in ['open youtube', 'youtube open']:
        webbrowser.open('www.youtube.com')
        speak('opening youtube')

    elif text in ['open telegram', 'telegram open']:
        speak('opening telegram')
        webbrowser.open('www.telegram.com')

    elif text in ["open resource monitor", "resource monitor open"]:
        speak("opening resource monitor")
        subprocess.run('''resmon''')

    elif text in ["open computer management console", "computer management console open", "open computer management", "computer management open"]:
        speak("opening computer management console")
        subprocess.run('''compmgmtlauncher''')

    elif text in ['open whatsapp', 'whatsapp open']:
        speak('opening whatsapp')
        webbrowser.open('www.whatsap.com')

    elif text in ['open gmail', 'gmail open']:
        speak('opening gmail')
        webbrowser.open('www.gmail.com')

    elif text in ['open github', 'github open']:
        speak('opening github')
        webbrowser.open('www.github.com')

    elif text in ['open linkedin', 'linkedin open']:
        speak('opening linkedin')
        webbrowser.open('www.linkedin.com')

    elif text in ['open instagram', 'instagram open']:
        speak('opening insta gram')
        webbrowser.open('www.instagram.com')

    elif text in ['open music', 'play music', 'music open', 'music play']:
        os.startfile("wmplayer")

    elif text in ["open documents", "documents open"]:
        speak("opening documents")
        os.startfile("Documents")

    elif text in ["open wordpad", "wordpad open"]:
        speak("opening wordpad")
        os.startfile("wordpad")

    elif text in ["open ms paint", "ms paint open", "open paint", "paint open", "open microsoft paint"]:
        speak("opening ms paint")
        os.startfile("mspaint")

    elif text in ["open calculator", "calculator open"]:
        speak("opening calculator")
        os.startfile("calc")

    elif text in ["open task manager", "task manager open"]:
        speak("opening task manager")
        os.startfile("taskmgr")

    elif text in ["open notepad", "notepad open"]:
        speak("opening notepad")
        os.startfile("notepad")

    elif text in ["open c drive", "open cdrive", "cdrive open", "c drive open"]:
        speak("opening c drive")
        os.startfile("c:")

    elif text in ["open settings", "settings open", "open setting", "setting open"]:
        speak("opening settings")
        os.startfile("ms-settings:")

    elif text in ["open microsoft store", "microsoft store open"]:
        speak("opening microsoft store")
        os.startfile("ms-windows-store:")

    elif text in ["check windows version", "windows version check", "check window version", "what is window version"]:
        speak("checking your windows version")
        os.startfile("winver")

    elif text in ["open phone dialler", "phone dialler open"]:
        speak("opening phone dialer")
        os.startfile("dialer")

    elif text in ["opens the temporary files folder", "open the temporary files folder", "open temporary files folder", "temporary files folder open", "open temporary files", "temporary file open", "open temporary file", "temporary files open"]:
        speak("opening the temporary files folder")
        os.startfile('''temp''')

    elif text in ["open camera", "camera open"]:
        speak("opening camera")
        os.startfile("microsoft.windows.camera:")

    elif text in ["Open the registry editor", "Opens the registry editor", "open registry editor", "resistry editor open"]:
        speak("Opening the Registry Editor")
        os.startfile("regedit")

    elif text in ["open disk management", "disk management open"]:
        speak("opening disk management")
        os.startfile("diskmgmt.msc")

    elif text in ["open this pc", "this pc open"]:
        speak("opening this pc")
        os.startfile('\"')

    elif text in ["open user account", "open user accounts", "user account open", "user accounts open"]:
        speak("opening user accounts")
        os.startfile("netplwiz")

    elif text in ["open control Panel", "control Panel open"]:
        speak("opening control Panel")
        os.startfile("control")

    elif text in ["open device manager", "device manager open"]:
        speak("opening device manager")
        os.startfile('''devmgmt.msc''')

    elif text in ["open power option", "power option open"]:
        speak("opening power option")
        os.startfile('''powercfg.cpl''')

    elif text in ["open the directx diagnostic tool", "directx diagnostic tool open", "open directx diagnostic tool", "directx diagnostic tool open", "open diagnostic tool", "diagnostic tool open"]:
        subprocess.run('''dxdiag''')
        speak("opening the directx diagnostic tool")

    elif text in ["open programs and features", "programs and features open", "open uninstall setting", "uninstall setting open"]:
        os.startfile('''appwiz.cpl''')
        speak("opening programs and features")

    elif text in ["open character map", "character map open"]:
        os.startfile('charmap')
        speak("opening character map")

    elif text in ["open network connections", "network connections open", "open network connection", 'network connection open']:
        speak("opening network connections")
        os.startfile("ncpa.cpl")

    elif text in ["open keyboard", "keyboard open"]:
        os.startfile("osk")

    elif text in ["open mouse properties", "mouse properties open"]:
        speak("opening mouse properties")
        os.startfile("main.cpl")

    elif text in ["open remote desktop connection", "remote desktop connection open"]:
        speak("opening remote desktop connection")
        os.startfile("mstsc")

    elif text in ["open file explorer", "file explorer open"]:
        os.startfile("explorer")
        speak("opening file explorer")

    elif text in ["shutdown", "shut down", "shutdown laptop", "laptop shut down", "laptop shutdown", "shut down laptop", "shutdown computer", "shut down computer", "computer shutdown", "computer shut down"]:
        a = subprocess.run("shutdown /s")
        speak("hello user i am going to shuting down your computer, wait some seconds. please don't touch any key until i shutdown computer")

    elif text in ["restart", "restart laptop", "laptop restart", "computer restart", "restart computer"]:
        a = subprocess.run("shutdown /r")
        speak("hello user i am going to shuting down your computer, wait some seconds. please don't touch any key until i restart computer")

    elif text in ["open powershell", "powershell open"]:
        os.startfile("powershell")
        speak("opening powershell")

    elif text in ["open excel", "excel open"]:
        os.startfile("excel")

    elif text in ["open chrome", "chrome open", "open google chrome", "google chrome open"]:
        os.startfile("chrome")
        speak("opening chrome")

    elif text in ["open brave", "brave open"]:
        os.startfile('Brave')
        speak("opening brave")

    elif text in ["open firefox", "firefox open"]:
        os.startfile('firefox')
        speak("opening firefox")

    elif text in ["open cmd", "cmd open", "open command prompt", "command prompt open"]:
        os.startfile("cmd")

    elif text in ["open code", "code open"]:
        os.startfile("code")

    elif text in ["wishme", "wish me"]:
        wishme(text)
    elif text in ["open microsoft edge", "Microsoft edge open"]:
        os.startfile("msedge")
        speak("opening microsoft edge")

    elif text in ["open photoshop", "photoshop open"]:
        os.startfile("photoshop")
        speak("opening photoshop")

    elif text in ["open outlook", "outlook open"]:
        os.startfile("outlook")
        speak("opening outlook")
        exit()
    elif text in ["open powerpoint", "powerpoint open"]:
        os.startfile("powerpnt")
        speak("opening powerpoint")

    elif text in ['how are you']:
        print(text)
        print()
        print('i am fine.\ntell me how may i help you')
        speak('i am fine, tell me how may i help you')
    elif text in ['what is your name', "what's your name"]:
        print(text)
        print()
        print('''well, my name's maxy" \ni wish that everyone\nhad a nickname as cool as mine\nso plz keep small and sort your name  ''')
        speak('''well, my name is maxy, i wish that everyone had a nickname as cool as mine, so plz keep small your name  ''')
    elif text in ['are you marry me', "will you marry me"]:
        print("this is one of things \nwe'd both have to agree\non i'd prefer to keep \nour friendship as it is.")
        speak("this is one of things, we'd both have to agree on i'd prefer to keep  our friendship as it is. ")
    elif text in ['what can you do for me']:
        print(text)
        print()
        print("i can do all the work \n which is in my might")
        speak("i can do all the work, which is in my might")
    elif text in ["do something for me"]:
        print(text)
        print()
        print("Ask me any problem \ni will try to solve it \nfor you")
        speak("Ask me any problem, i will try to solve it for you")
    elif text in ['date', "what's date", "what is date", "date", "what's the date today", "today date", "today's date", "what is the date", "what's the date"]:
        print(text)
        print()
        print(date)
        speak(date)
    elif text in ["tell me some jokes", "tell some jokes", "tell me some joke", "kucch joke sunao", "kuchh jokes sunao", 'tell me joke ', 'tell me jokes']:
        print(text)
        print()
        print("Air hostess asked lalu \nPrasad yadav. \nSir are you vegetarian or \nNon vegetarian \nLalu said I am indian \nAir hostess said okay, \nAre you shakahari or mansahari \nLalu said hat sasuri I am Bihari")
        speak("Air hostess asked lalu Prasad yadav. Sir are you vegetarian or Non vegetarian, Lalu said I am indian. Air hostess said okay, Are you shakahari or mansahari, Lalu said hat sasuri I am Bihari")
    elif "wikipedia" in text:
        result = wikipedia.summary(text, sentences=2)
        print(result)
        speak(result)
    elif "print table of" in text:
        print(text)

        nu = text.split()
        nu = int(nu[-1])
        for i in range(1, 11):
            print(i*nu, end=" ")
        print()

    elif "on youtube" in text:
        try:
            wtp.playonyt(text)
            speak('playing')
        except:
            speak("network Error Occurred ")

    elif "sum of" in text or "add" in text:
        print()
        print(text)
        print()
        text = text.split()
        li = []
        sum = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                sum += float(i)

        print("the answer is", sum)
        speak("the answer is "+str(sum))
    elif "area of circle" in text:
        text = text.split()
        li = []
        rad = 0
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                rad = float(i)
        area = 3.14*rad*rad
        print("The area of circle is ", area)
        speak("The area of circle is "+str(area))

    elif "multiply" in text:
        print()
        print(text)
        print()
        text = text.split()
        multp = 1
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                multp *= float(i)
        print("the answer is", multp)
        speak("the answer is "+str(multp))
    elif ("divided" in text) or ("/" in text):
        print()
        print(text)
        print()
        text = text.split()
        li = []
        for i in text:
            if i[0] >= '0' and i[0] <= '9':
                li.append(float(i))
        a = li[0]
        b = li[1]
        print("the answer is", a/b)
        speak("the answer is"+str(a/b))

    elif "how to make" in text:
        try:
            wtp.playonyt(text)
            speak("playing")
        except:
            speak("network Error Occurred ")

    elif text in ["do you know chitkara university"]:
        print(text)
        print()
        print("yes i know chitkara university, \nit is the  best private university in the punjab ")
        speak(
            "yes i know chitkara university, it is best private university in the punjab ")
    elif "factorial" in text:
        print(text)
        print()

        fact = str(text)
        fact = fact.split()
        fact = int(fact[-1])

        fact = math.factorial(fact)
        print(fact)
        speak(fact)
    elif text in ["open coding ninjas", "coding ninjas open", "coding ninjas"]:
        webbrowser.open('https://www.codingninjas.com')

    elif text in ["open vs code", "open visual studio code", "vs code open", "visual studio code open"]:
        vs_pasth = "C:\\Users\\lalit\\OneDrive\\Desktop\\Visual Studio Code.lnk"
        webbrowser.open(vs_pasth)

    else:
        print(text)
        print("sorry i don't understand")
        speak("sorry i don't understand")


if __name__ == '__main__':
    while(True):
        text = command()

        doooo(text)
